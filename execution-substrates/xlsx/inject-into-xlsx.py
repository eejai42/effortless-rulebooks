#!/usr/bin/env python3
"""
General-purpose Rulebook-to-XLSX transpiler.

This script reads the effortless-rulebook.json and generates an Excel workbook
with one worksheet per table, including:
- All columns from the schema
- Data rows with raw values
- Excel formulas for calculated fields

The script is generic and works for ANY rulebook structure - it simply reads
what's defined and generates the corresponding xlsx.

Smart Update Feature:
To avoid unnecessary git dirty state from volatile metadata (e.g., file
timestamps), this script compares the newly generated xlsx content (data +
formulas) against git HEAD. If they're identical, it restores the git version
via `git checkout`, preventing spurious uncommitted changes.
"""

import sys
import re
from pathlib import Path

# Add project root to path for shared imports
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from orchestration.shared import load_rulebook, get_candidate_name_from_cwd, handle_clean_arg

# Try to import openpyxl, provide helpful error if missing
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def get_table_names(rulebook):
    """Extract table names from the rulebook (excluding metadata keys)."""
    metadata_keys = {'$schema', 'model_name', 'Description', '_meta'}
    return [key for key in rulebook.keys() if key not in metadata_keys]


def build_column_map(schema):
    """Build a mapping of field names to column letters.

    Returns a dict like {'Name': 'A', 'Category': 'B', ...}
    """
    column_map = {}
    for idx, field in enumerate(schema):
        col_letter = get_column_letter(idx + 1)
        column_map[field['name']] = col_letter
    return column_map


def convert_formula_to_excel(formula, column_map, row_num):
    """Convert a rulebook formula to an Excel formula.

    Converts {{FieldName}} placeholders to cell references like $B2.
    The $ before the column letter makes it an absolute column reference,
    while the row number remains relative.

    Args:
        formula: The formula string from the rulebook (e.g., "={{HasSyntax}} = TRUE()")
        column_map: Dict mapping field names to column letters
        row_num: The current row number (1-indexed, accounting for header row)

    Returns:
        Excel formula string with cell references
    """
    result = formula

    # Find all {{FieldName}} patterns and replace with cell references
    pattern = r'\{\{(\w+)\}\}'

    def replace_field(match):
        field_name = match.group(1)
        if field_name in column_map:
            col_letter = column_map[field_name]
            return f'${col_letter}{row_num}'
        else:
            # Keep the placeholder if field not found (might be an error in rulebook)
            return match.group(0)

    result = re.sub(pattern, replace_field, result)
    return result


def evaluate_formula(formula, row_data):
    """Evaluate a rulebook formula using row data.

    Handles these formula patterns:
    - String concatenation: ="Is " & {{Name}} & " a language?"
    - AND: =AND({{A}}, {{B}}, NOT({{C}}), {{D}}=2)
    - IF: =IF(condition, true_val, false_val)
    - NOT: =NOT({{Field}})
    - Equality: ={{Field}} = TRUE()

    Args:
        formula: The formula string from the rulebook
        row_data: Dict of field values for the current row

    Returns:
        The computed value
    """
    if not formula.startswith('='):
        return formula

    def get_field(name):
        """Get field value from row_data."""
        return row_data.get(name)

    def normalize_ampersands(s):
        """Normalize & operators to have consistent spacing, respecting strings."""
        result = []
        in_string = False
        i = 0
        while i < len(s):
            c = s[i]
            if c == '"':
                in_string = not in_string
                result.append(c)
            elif c == '&' and not in_string:
                # Ensure space before & (if not already there)
                if result and result[-1] not in (' ', ''):
                    result.append(' ')
                result.append('&')
                # Skip any existing whitespace after &
                i += 1
                while i < len(s) and s[i] == ' ':
                    i += 1
                # Add single space after &
                result.append(' ')
                continue  # i already incremented past whitespace
            else:
                result.append(c)
            i += 1
        return ''.join(result)

    def eval_expr(expr):
        """Recursively evaluate an expression."""
        expr = expr.strip()

        # Normalize & spacing to handle variations like &IF vs & IF
        expr = normalize_ampersands(expr)

        # Handle string concatenation (only if & is at top level)
        parts = split_by_operator(expr, ' & ')
        if len(parts) > 1:
            result = ''
            for p in parts:
                val = eval_expr(p)
                if val is not None:
                    result += str(val)
            return result if result else None

        # Handle string literals
        if expr.startswith('"') and expr.endswith('"'):
            return expr[1:-1]

        # Handle TRUE/FALSE
        if expr.upper() in ('TRUE', 'TRUE()'):
            return True
        if expr.upper() in ('FALSE', 'FALSE()'):
            return False

        # Handle field references {{FieldName}}
        field_match = re.match(r'^\{\{(\w+)\}\}$', expr)
        if field_match:
            return get_field(field_match.group(1))

        # Handle numeric literals
        try:
            if '.' in expr:
                return float(expr)
            return int(expr)
        except ValueError:
            pass

        # Handle AND(...)
        if expr.upper().startswith('AND('):
            inner = extract_parens(expr[3:])
            args = split_args(inner)
            for arg in args:
                val = eval_expr(arg)
                if not val:
                    return False
            return True

        # Handle OR(...)
        if expr.upper().startswith('OR('):
            inner = extract_parens(expr[2:])
            args = split_args(inner)
            for arg in args:
                val = eval_expr(arg)
                if val:
                    return True
            return False

        # Handle NOT(...)
        if expr.upper().startswith('NOT('):
            inner = extract_parens(expr[3:])
            val = eval_expr(inner)
            return not val if val is not None else None

        # Handle IF(...)
        if expr.upper().startswith('IF('):
            inner = extract_parens(expr[2:])
            args = split_args(inner)
            if len(args) < 2:
                return None
            condition = eval_expr(args[0])
            true_val = eval_expr(args[1]) if len(args) > 1 else None
            false_val = eval_expr(args[2]) if len(args) > 2 else None
            return true_val if condition else false_val

        # Handle SUM(...)
        if expr.upper().startswith('SUM('):
            inner = extract_parens(expr[3:])
            args = split_args(inner)
            total = 0
            for arg in args:
                val = eval_expr(arg)
                if val is not None:
                    try:
                        total += float(val) if isinstance(val, (int, float)) else float(val)
                    except (ValueError, TypeError):
                        pass
            return int(total) if total == int(total) else total

        # Handle LOWER(...)
        if expr.upper().startswith('LOWER('):
            inner = extract_parens(expr[5:])
            val = eval_expr(inner)
            if val is None:
                return ''
            return str(val).lower()

        # Handle UPPER(...)
        if expr.upper().startswith('UPPER('):
            inner = extract_parens(expr[5:])
            val = eval_expr(inner)
            if val is None:
                return ''
            return str(val).upper()

        # Handle SUBSTITUTE(text, old, new)
        if expr.upper().startswith('SUBSTITUTE('):
            inner = extract_parens(expr[10:])
            args = split_args(inner)
            if len(args) < 3:
                return None
            text = eval_expr(args[0])
            old_text = eval_expr(args[1])
            new_text = eval_expr(args[2])
            if text is None:
                return ''
            old_str = str(old_text) if old_text is not None else ''
            new_str = str(new_text) if new_text is not None else ''
            return str(text).replace(old_str, new_str)

        # Handle comparison operators (check multi-char operators first)
        for op, op_fn in [(' >= ', lambda a, b: a >= b),
                          (' <= ', lambda a, b: a <= b),
                          (' > ', lambda a, b: a > b),
                          (' < ', lambda a, b: a < b),
                          ('>=', lambda a, b: a >= b),
                          ('<=', lambda a, b: a <= b),
                          ('>', lambda a, b: a > b),
                          ('<', lambda a, b: a < b)]:
            if op in expr:
                parts = expr.split(op, 1)
                if len(parts) == 2:
                    left = eval_expr(parts[0])
                    right = eval_expr(parts[1])
                    if left is not None and right is not None:
                        return op_fn(left, right)

        # Handle equality: {{Field}} = value or value = value
        if ' = ' in expr or '=' in expr:
            # Split by = but be careful of ==
            parts = re.split(r'\s*=\s*', expr, maxsplit=1)
            if len(parts) == 2:
                left = eval_expr(parts[0])
                right = eval_expr(parts[1])
                return left == right

        return None

    def extract_parens(s):
        """Extract content inside parentheses."""
        s = s.strip()
        if not s.startswith('('):
            return s
        depth = 0
        for i, c in enumerate(s):
            if c == '(':
                depth += 1
            elif c == ')':
                depth -= 1
                if depth == 0:
                    return s[1:i]
        return s[1:-1] if s.endswith(')') else s

    def split_by_operator(expr, op):
        """Split by operator, respecting parentheses and quotes."""
        parts = []
        current = ''
        depth = 0
        in_string = False
        i = 0
        while i < len(expr):
            c = expr[i]
            if c == '"':
                in_string = not in_string
                current += c
            elif not in_string:
                if c == '(':
                    depth += 1
                    current += c
                elif c == ')':
                    depth -= 1
                    current += c
                elif depth == 0 and expr[i:i+len(op)] == op:
                    parts.append(current)
                    current = ''
                    i += len(op) - 1
                else:
                    current += c
            else:
                current += c
            i += 1
        if current:
            parts.append(current)
        return parts

    def split_args(s):
        """Split comma-separated arguments, respecting parens and quotes."""
        return split_by_operator(s, ',')

    return eval_expr(formula[1:])


def get_value_for_cell(field_schema, row_data, column_map, row_num):
    """Get the value to put in a cell.

    For raw fields, returns the data value.
    For calculated fields with formulas, returns the Excel formula.

    Args:
        field_schema: The field definition from the schema
        row_data: The data row dict
        column_map: Dict mapping field names to column letters
        row_num: The current row number (1-indexed)

    Returns:
        The value or Excel formula to put in the cell
    """
    field_name = field_schema['name']
    field_type = field_schema.get('type', 'raw')

    if field_type == 'calculated' and 'formula' in field_schema:
        # This is a calculated field - use the Excel formula
        formula = field_schema['formula']
        return convert_formula_to_excel(formula, column_map, row_num)
    else:
        # This is a raw field - use the data value
        value = row_data.get(field_name)

        # Handle special cases
        if value is None:
            return ''
        elif isinstance(value, bool):
            return value
        else:
            return value


def apply_header_style(cell):
    """Apply styling to header cells."""
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def apply_data_style(cell, is_calculated=False):
    """Apply styling to data cells."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border

    if is_calculated:
        # Light blue background for calculated fields
        cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')


def create_worksheet_from_table(workbook, table_name, table_data):
    """Create a worksheet from a rulebook table definition.

    Args:
        workbook: The openpyxl Workbook
        table_name: Name of the table (used as sheet name)
        table_data: Dict with 'schema', 'data', and optionally 'Description'

    Returns:
        The created worksheet
    """
    ws = workbook.create_sheet(title=table_name[:31])  # Excel limits sheet names to 31 chars

    schema = table_data.get('schema', [])
    data = table_data.get('data', [])

    if not schema:
        return ws

    # Build column map for formula conversion
    column_map = build_column_map(schema)

    # Create a map of field names to their type (raw vs calculated)
    field_types = {f['name']: f.get('type', 'raw') for f in schema}

    # Write header row
    for col_idx, field in enumerate(schema, 1):
        cell = ws.cell(row=1, column=col_idx, value=field['name'])
        apply_header_style(cell)

        # Add description as cell comment if available
        description = field.get('Description', '')
        if description:
            cell.comment = Comment(description, "ERB Rulebook")

        # Set column width based on header length (minimum 12 chars)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(field['name']) + 2, 12)

    # Write data rows
    for row_idx, row_data in enumerate(data, 2):  # Start at row 2 (after header)
        for col_idx, field in enumerate(schema, 1):
            value = get_value_for_cell(field, row_data, column_map, row_idx)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_calculated = field.get('type') == 'calculated'
            apply_data_style(cell, is_calculated)

    # Freeze the header row
    ws.freeze_panes = 'A2'

    return ws


def export_xlsx_to_json(xlsx_path):
    """Export xlsx content to a JSON-serializable structure for comparison.

    Exports data and formulas (not volatile metadata like timestamps).
    This allows comparing two xlsx files for meaningful content changes.

    Args:
        xlsx_path: Path to the xlsx file

    Returns:
        Dict with sheet data, or None on error
    """
    try:
        wb = load_workbook(xlsx_path)
        result = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = []

            for row in ws.iter_rows():
                row_data = []
                for cell in row:
                    # Get the cell value (formula string if it's a formula, else value)
                    if cell.value is None:
                        row_data.append(None)
                    elif isinstance(cell.value, str) and cell.value.startswith('='):
                        # It's a formula - store the formula itself
                        row_data.append(cell.value)
                    else:
                        row_data.append(cell.value)
                sheet_data.append(row_data)

            result[sheet_name] = sheet_data

        wb.close()
        return result

    except Exception as e:
        print(f"  Error exporting xlsx to JSON: {e}")
        return None


def get_git_file_content(file_path):
    """Get the git HEAD version of a file as bytes.

    Args:
        file_path: Path to the file (relative or absolute)

    Returns:
        Bytes content of the file from git HEAD, or None if not in git/doesn't exist
    """
    import subprocess

    try:
        # Get the relative path from the git root
        result = subprocess.run(
            ['git', 'rev-parse', '--show-toplevel'],
            capture_output=True, text=True, check=True
        )
        git_root = Path(result.stdout.strip())

        # Make path relative to git root
        abs_path = Path(file_path).resolve()
        try:
            rel_path = abs_path.relative_to(git_root)
        except ValueError:
            print(f"  File {file_path} is not under git root")
            return None

        # Get file content from HEAD
        result = subprocess.run(
            ['git', 'show', f'HEAD:{rel_path}'],
            capture_output=True, check=True
        )
        return result.stdout

    except subprocess.CalledProcessError:
        # File doesn't exist in git or not a git repo
        return None
    except Exception as e:
        print(f"  Error getting git file: {e}")
        return None


def export_git_xlsx_to_json(file_path):
    """Export the git HEAD version of an xlsx file to JSON for comparison.

    Args:
        file_path: Path to the xlsx file

    Returns:
        Dict with sheet data, or None if file not in git or on error
    """
    import tempfile

    git_content = get_git_file_content(file_path)
    if git_content is None:
        return None

    # Write to temp file and load with openpyxl
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(git_content)
            tmp_path = tmp.name

        result = export_xlsx_to_json(tmp_path)

        # Clean up temp file
        Path(tmp_path).unlink()

        return result

    except Exception as e:
        print(f"  Error loading git xlsx: {e}")
        return None


def git_checkout_file(file_path):
    """Restore a file to its git HEAD version.

    Args:
        file_path: Path to the file

    Returns:
        True if successful, False otherwise
    """
    import subprocess

    try:
        subprocess.run(
            ['git', 'checkout', 'HEAD', '--', str(file_path)],
            capture_output=True, check=True
        )
        return True
    except subprocess.CalledProcessError as e:
        print(f"  Error checking out file: {e}")
        return False


def cleanup_file(path):
    """Safely remove a file if it exists."""
    path = Path(path)
    if path.exists():
        path.unlink()


def generate_workbook(rulebook, table_names):
    """Generate the workbook from the rulebook.
    
    Args:
        rulebook: The loaded rulebook dict
        table_names: List of table names to process
        
    Returns:
        The generated Workbook object
    """
    wb = Workbook()
    default_sheet = wb.active

    for table_name in table_names:
        table_data = rulebook[table_name]

        if not isinstance(table_data, dict) or 'schema' not in table_data:
            print(f"  Skipping {table_name}: not a table structure")
            continue

        print(f"  Creating worksheet: {table_name}")
        schema = table_data.get('schema', [])
        data = table_data.get('data', [])

        raw_count = sum(1 for f in schema if f.get('type', 'raw') == 'raw')
        calc_count = sum(1 for f in schema if f.get('type') == 'calculated')

        print(f"    - {len(schema)} columns ({raw_count} raw, {calc_count} calculated)")
        print(f"    - {len(data)} data rows")

        create_worksheet_from_table(wb, table_name, table_data)

    if len(wb.sheetnames) > 1:
        wb.remove(default_sheet)

    return wb


def main():
    # Define generated files for this substrate
    GENERATED_FILES = [
        'rulebook.xlsx',
        'test-answers.json',
        'test-results.md',
    ]

    # Handle --clean argument
    if handle_clean_arg(GENERATED_FILES, "XLSX substrate: Removes generated Excel workbook and test outputs"):
        return

    candidate_name = get_candidate_name_from_cwd()
    print(f"Generating {candidate_name} from rulebook...")

    # Define paths
    output_path = Path('rulebook.xlsx')

    # Load the rulebook
    try:
        rulebook = load_rulebook()
    except FileNotFoundError as e:
        print(f"Error: {e}")
        sys.exit(1)

    # Get all table names from the rulebook
    table_names = get_table_names(rulebook)

    if not table_names:
        print("Warning: No tables found in rulebook")
        sys.exit(1)

    print(f"Found {len(table_names)} tables: {', '.join(table_names)}")

    # Step 1: Get git's version of the file BEFORE we overwrite it
    print(f"\n--- Smart Update: Getting git baseline ---")
    git_content = export_git_xlsx_to_json(output_path)
    if git_content:
        print(f"  Exported git HEAD version for comparison")
    else:
        print(f"  No git baseline (new file or not in git)")

    # Step 2: Generate the new workbook (always)
    print(f"\nGenerating new xlsx...")
    wb = generate_workbook(rulebook, table_names)

    # Save the new workbook
    wb.save(output_path)
    print(f"\nGenerated: {output_path}")
    print(f"  - {len(wb.sheetnames)} worksheets")

    # Step 3: Compare with git and restore if identical
    if git_content:
        print(f"\n--- Smart Update: Comparing with git ---")
        new_content = export_xlsx_to_json(output_path)

        if new_content == git_content:
            # Content is identical - restore git's version to avoid dirty state
            print(f"  NO CONTENT CHANGE - restoring git version")
            if git_checkout_file(output_path):
                print(f"  Restored {output_path} from git HEAD")
                print(f"\n*** XLSX NOT UPDATED (content unchanged) ***")
            else:
                print(f"  Warning: Could not restore git version")
        else:
            print(f"  CONTENT CHANGED - keeping new xlsx")

    print(f"\nDone generating {candidate_name}.")


if __name__ == "__main__":
    main()
