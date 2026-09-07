#!/usr/bin/env python3
# =============================================================================
# TEST ORCHESTRATOR (Generic / Domain-Agnostic)
# =============================================================================
# Generic test framework for evaluating execution substrates.
#
# This orchestrator knows NOTHING about specific domains. It only knows:
# - Views (anything named vw_* in postgres)
# - Raw fields (schema type: "raw")
# - Computed fields (schema type: "calculated" or "aggregation")
# - JSON comparison (expected vs actual)
#
# All configuration is auto-discovered from the rulebook and database.
# =============================================================================

import glob
import json
import os
import re
import subprocess
import sys
import time
from datetime import datetime

import psycopg2
from psycopg2.extras import RealDictCursor

# The formula parser is a hard dependency of answer-key generation. If it
# can't be imported, this entire process is bogus — we cannot validate any
# substrate's output. Let the ImportError propagate.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from formula_parser import evaluate_field

# =============================================================================
# PATHS (Generic - No Domain-Specific Names)
# =============================================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)
PROJECT_ROOT = REPO_ROOT  # kept for backward-compat; prefer REPO_ROOT below
SUBSTRATES_DIR = os.path.join(REPO_ROOT, "execution-substrates")
SUMMARY_PATH = os.path.join(SCRIPT_DIR, "all-tests-results.md")

def _get_active_domain():
    domain = os.environ.get("ERB_DOMAIN", "").strip()
    if not domain:
        raise RuntimeError(
            "ERB_DOMAIN is not set. test-orchestrator must be invoked with "
            "ERB_DOMAIN=<slug> in its environment (orchestrate.sh does this "
            "automatically for menu-driven runs)."
        )
    return domain

def _find_domain_dir(domain):
    # Mirrors orchestrate.sh's find_domain_dir(): check rulebook-examples/
    # first, then toy-rulebooks/. Fails loudly if the domain is in neither.
    examples_dir = os.path.join(REPO_ROOT, "rulebook-examples", domain)
    toy_dir = os.path.join(REPO_ROOT, "toy-rulebooks", domain)
    if os.path.isdir(examples_dir):
        return examples_dir
    if os.path.isdir(toy_dir):
        return toy_dir
    raise RuntimeError(
        f"Domain '{domain}' not found under rulebook-examples/ or toy-rulebooks/ "
        f"(checked {examples_dir} and {toy_dir})."
    )

ACTIVE_DOMAIN = _get_active_domain()
DOMAIN_DIR = _find_domain_dir(ACTIVE_DOMAIN)
RULEBOOK_DIR = os.path.join(DOMAIN_DIR, "effortless-rulebook")
RULEBOOK_PATH = os.path.join(RULEBOOK_DIR, f"{ACTIVE_DOMAIN}-rulebook.json")

# All conformance artifacts live inside the domain folder so each rulebook
# example is fully self-contained. The central testing/ folder at repo root
# is no longer used.
TESTING_DIR = os.path.join(DOMAIN_DIR, "testing")
ANSWER_KEYS_DIR = os.path.join(TESTING_DIR, "answer-keys")
BLANK_TESTS_DIR = os.path.join(TESTING_DIR, "blank-tests")

# Rulebook tables that are NOT graded by conformance. `discover_entities()`
# (shared.py) deliberately INCLUDES `__meta__` because injectors materialize it
# like any other table (CSV sheet, Airtable table, etc.). The Postgres
# transpiler, however, intentionally emits NO `vw___meta__` view — `__meta__` is
# project-level metadata, not domain data (see CLAUDE.md's `__meta__` doctrine).
# Grading it would therefore score every row as a failure against a view that
# does not (and should not) exist. We exclude it here, at the one place the
# entity list feeds answer-key generation, so it propagates to blank-tests,
# _metadata.json, and grading in a single spot.
NON_GRADED_ENTITIES = {"__meta__"}

def get_substrate_test_answers_dir(substrate_name: str) -> str:
    """Return the domain-scoped test-answers dir for a substrate."""
    return os.path.join(TESTING_DIR, substrate_name, "test-answers")

# Canonical substrate ordering — matches SUBSTRATE_ORDER in orchestrate.sh.
# Substrates not in this list fall through to the end, alphabetically.
SUBSTRATE_ORDER = [
    "english",
    "python",
    "golang",
    "owl",
    "uml",
    "xlsx",
    "binary",
    "cobol",
    "csv",
    "yaml",
    "explain-dag",
    # Effortless-licensed substrates render LAST in the report — they use
    # the production rulebook-to-X transpiler pipelines and should always
    # be 100% conformant.
    "effortless-postgres",
    "effortless-xlsx",
    "effortless-entity-framework",
]

# Database connection — DATABASE_URL overrides; otherwise default to the
# active domain's per-domain DB (erb_<domain>), derived from ERB_DOMAIN.
DB_CONNECTION = os.environ.get("DATABASE_URL") or (
    "postgresql://postgres@localhost:5432/erb_" + ACTIVE_DOMAIN.replace("-", "_")
)

# =============================================================================
# ANSI Color Codes (Unchanged)
# =============================================================================

GREEN = "\033[92m"
RED = "\033[91m"
YELLOW = "\033[93m"
RESET = "\033[0m"
BOLD = "\033[1m"
SKY_BLUE_BG = "\033[48;5;117m"
DARK_TEXT = "\033[38;5;232m"
GREEN_BG = "\033[48;5;22m"
RED_BG = "\033[48;5;52m"
WHITE_TEXT = "\033[97m"
STRIKETHROUGH = "\033[9m"
DIM = "\033[2m"


def get_score_color(score: float) -> str:
    """Returns ANSI color code for a score using a red->yellow->green gradient."""
    if score >= 100:
        return "\033[38;5;46m"
    elif score >= 90:
        return "\033[38;5;82m"
    elif score >= 80:
        return "\033[38;5;118m"
    elif score >= 70:
        return "\033[38;5;154m"
    elif score >= 60:
        return "\033[38;5;190m"
    elif score >= 50:
        return "\033[38;5;226m"
    elif score >= 40:
        return "\033[38;5;220m"
    elif score >= 30:
        return "\033[38;5;214m"
    elif score >= 20:
        return "\033[38;5;208m"
    elif score >= 10:
        return "\033[38;5;202m"
    else:
        return "\033[38;5;196m"


# =============================================================================
# RUN METADATA: Track success/failure history per substrate
# =============================================================================

# Central substrate results file - THE source of truth for all substrate stats
CENTRAL_RESULTS_PATH = os.path.join(TESTING_DIR, "_substrate_results.json")


def load_central_results() -> dict:
    """Load the central _substrate_results.json from testing folder"""
    if os.path.exists(CENTRAL_RESULTS_PATH):
        with open(CENTRAL_RESULTS_PATH, 'r') as f:
            return json.load(f)
    return {}


def save_central_results(results: dict):
    """Save the central _substrate_results.json to testing folder"""
    with open(CENTRAL_RESULTS_PATH, 'w') as f:
        json.dump(results, f, indent=2, default=str)


def load_run_metadata(substrate_name: str) -> dict:
    """Load metadata for a substrate from the CENTRAL results file"""
    central = load_central_results()
    return central.get(substrate_name, {"last_run": None, "last_successful_run": None})


def save_run_metadata(substrate_name: str, metadata: dict):
    """Save metadata for a substrate to the CENTRAL results file"""
    central = load_central_results()
    central[substrate_name] = metadata
    save_central_results(central)


def update_run_metadata(substrate_name: str, grades: dict, success: bool, error_msg: str = None, preserve_timing: bool = False):
    """Update run metadata after a test run.

    IMPORTANT: Only updates last_successful_run when score is 100%.
    This preserves previous successful run data including duration.
    If substrate was skipped, does not update metadata at all.

    Args:
        substrate_name: Name of the substrate
        grades: Grading results dict
        success: Whether the take-test.sh returned exit code 0
        error_msg: Error message if run failed
        preserve_timing: If True, preserve timing from previous run (user skipped LLM but still graded)
    """
    # If substrate was skipped by user, don't update metadata
    if grades.get("skipped"):
        return

    metadata = load_run_metadata(substrate_name)
    timestamp = datetime.now().isoformat()

    # When preserve_timing is True, use the timing already in grades (which was loaded from previous run)
    # Otherwise use the newly measured elapsed time
    elapsed = grades.get("elapsed_seconds", 0.0)

    total = grades.get("total_fields_tested", 0)
    passed = grades.get("fields_passed", 0)
    failed = grades.get("fields_failed", 0)
    score = (passed / total * 100) if total > 0 else 0.0

    # When preserving timing, we still update last_run but with preserved duration
    # This allows the test results to update while keeping the meaningful timing
    run_record = {
        "status": "success" if success else "failure",
        "duration_seconds": elapsed,
        "exit_code": 0 if success else 1,
        "score": score
    }
    if error_msg:
        run_record["error_message"] = error_msg

    metadata["last_run"] = run_record

    # Only update last_successful_run on success AND 100% score
    # When preserve_timing is True and score is 100%, we update the test_results
    # but keep the duration from the previous successful run
    if success and total > 0 and score >= 100.0:
        if preserve_timing and metadata.get("last_successful_run"):
            # Keep existing duration, just update test results and timestamp
            prev_duration = metadata["last_successful_run"].get("duration_seconds", elapsed)
            metadata["last_successful_run"] = {
                "duration_seconds": prev_duration,
                "status": "success",
                "test_results": {
                    "total_fields_tested": total,
                    "fields_passed": passed,
                    "fields_failed": failed,
                    "score": score
                }
            }
        else:
            metadata["last_successful_run"] = {
                "duration_seconds": elapsed,
                "status": "success",
                "test_results": {
                    "total_fields_tested": total,
                    "fields_passed": passed,
                    "fields_failed": failed,
                    "score": score
                }
            }

    save_run_metadata(substrate_name, metadata)


# =============================================================================
# AUTO-DISCOVERY: Views, Computed Columns, Primary Keys
# =============================================================================

def load_rulebook() -> dict:
    """Load the active domain's <domain>-rulebook.json file. Fails loudly if missing."""
    if not os.path.exists(RULEBOOK_PATH):
        raise FileNotFoundError(
            f"Rulebook not found at {RULEBOOK_PATH}. "
            f"Active domain is '{ACTIVE_DOMAIN}' — its rulebook MUST be at this exact path. "
            "Fix the file/name rather than substituting a different rulebook."
        )
    with open(RULEBOOK_PATH, 'r') as f:
        return json.load(f)


# Schema-introspection helpers are defined ONCE in shared.py. Importing them
# here keeps test-orchestrator and every other orchestration script in sync.
from shared import (
    to_snake_case,
    to_pascal_case,
    discover_entities,
    get_entity_schema,
    get_entity_data,
    discover_primary_key,
    discover_computed_columns,
)


def discover_views(conn) -> list:
    """Query postgres for all vw_* views"""
    cur = conn.cursor()
    cur.execute("""
        SELECT table_name
        FROM information_schema.views
        WHERE table_name LIKE 'vw_%'
        ORDER BY table_name
    """)
    views = [row[0] for row in cur.fetchall()]
    cur.close()
    return views


def view_to_entity_name(view_name: str) -> str:
    """Convert view name to entity name: vw_products -> products"""
    return view_name.replace('vw_', '')


# =============================================================================
# STEP 1: Generate Answer Keys from Rulebook
# =============================================================================
# Answer keys are generated directly from the rulebook's seed data.
# The rulebook is the single source of truth - no database required.
# All substrates (including Postgres) are tested against these answer keys.
# (get_entity_data is imported from shared.py above.)


def convert_record_to_snake_case(record: dict) -> dict:
    """Convert all keys in a record from PascalCase to snake_case."""
    return {to_snake_case(k): v for k, v in record.items()}


def coerce_record_types(record: dict, schema: list) -> dict:
    """Coerce record values to match their schema datatypes.

    - Integer/number fields: string numerics are parsed to numbers.
    - Boolean fields: strings "true"/"false" and empty/null are canonicalized
      to True/False. Airtable's lookup of an unchecked checkbox arrives as
      an empty string, so boolean blanks must normalize to False here — not
      treating them as booleans leaves the answer key out of sync with every
      substrate that correctly computes False.
    """
    type_map = {}
    for field in schema:
        field_name = to_snake_case(field.get('name', ''))
        datatype = field.get('datatype', 'string')
        type_map[field_name] = datatype

    coerced = {}
    for key, value in record.items():
        datatype = type_map.get(key, 'string')

        if datatype == 'boolean':
            # Canonical: empty/null/"false"/0 -> False; "true"/1/True -> True
            if value is None or value == '':
                coerced[key] = False
            elif isinstance(value, bool):
                coerced[key] = value
            elif isinstance(value, str):
                coerced[key] = value.strip().lower() in ('true', '1', 'yes')
            elif isinstance(value, (int, float)):
                coerced[key] = bool(value)
            else:
                coerced[key] = value
            continue

        if value is None or value == '':
            coerced[key] = value
            continue

        if datatype == 'integer':
            if isinstance(value, str):
                try:
                    coerced[key] = int(value)
                except ValueError:
                    coerced[key] = value
            else:
                coerced[key] = value
        elif datatype == 'number':
            if isinstance(value, str):
                try:
                    coerced[key] = float(value)
                except ValueError:
                    coerced[key] = value
            else:
                coerced[key] = value
        else:
            coerced[key] = value

    return coerced


def _recompute_calculated_fields(
    record: dict,
    schema: list,
    entity_pascal: str,
    failures: list,
) -> dict:
    """Re-evaluate scalar `type=calculated` formulas against `record`.

    The rulebook's seed `data` can drift from the current `formula` definitions
    (e.g. an Airtable export captures values from a previous formula version).
    The answer-keys must reflect the CURRENT formulas, so we evaluate them here.

    Per-field evaluation failures are appended to `failures` (with full context:
    entity, field, formula, error) and the caller decides what to do at the
    end. We continue evaluating other fields so the user sees ALL the failures
    from a single broken record in one report, not just the first one.
    """
    out = dict(record)
    pk_name = None
    for f in schema:
        if f.get('nullable') is False:
            pk_name = to_snake_case(f['name'])
            break
    record_id = out.get(pk_name) if pk_name else None

    for field in schema:
        if field.get('type') != 'calculated':
            continue
        formula = field.get('formula')
        if not formula:
            continue
        snake = to_snake_case(field['name'])
        stored = out.get(snake)
        has_stored = stored is not None

        try:
            value = evaluate_field(formula, out)
        except Exception as e:
            # The Python formula engine could not evaluate this formula. That is
            # a limitation of THIS engine (an unimplemented function, or a
            # time-dependent NOW() it can't reproduce against the substrate's
            # seeded clock) — NOT evidence the rulebook is unhealthy. The SSoT
            # oracle is the substrate (Postgres), whose value was already adopted
            # into the rulebook by regenerate-answer-keys.sh.
            #   • If we HAVE that substrate value, keep it and move on silently.
            #   • If we DON'T, there is nothing to fall back to — record it so the
            #     report surfaces a genuinely missing answer (no silent blank).
            if has_stored:
                continue
            failures.append({
                'entity': entity_pascal,
                'field': field['name'],
                'formula': formula,
                'record_id': record_id,
                'error': f"{type(e).__name__}: {e}",
                'severity': 'missing-value',
            })
            continue

        # The substrate-adopted stored value is authoritative — it is the answer
        # key's source of truth (CLAUDE.md: "the view IS the contract"). The
        # Python recompute is a CROSS-CHECK, not a second oracle: when it
        # disagrees with the stored value, surface drift (the stored seed is
        # stale vs the current formula → re-run regenerate-answer-keys.sh) but do
        # NOT overwrite the substrate's answer with the Python engine's.
        if has_stored:
            if value is not None and _norm_cmp(value) != _norm_cmp(stored):
                failures.append({
                    'entity': entity_pascal,
                    'field': field['name'],
                    'formula': formula,
                    'record_id': record_id,
                    'error': f"drift: stored={stored!r} python-recompute={value!r}",
                    'severity': 'drift',
                })
            # Keep the stored (substrate) value regardless.
            continue

        # No stored value: the Python recompute is the only source we have, so
        # adopt it (the original behavior for un-seeded calc fields).
        if value is not None:
            out[snake] = value
    return out


def _norm_cmp(v):
    """Loose equality normalizer for drift comparison (bool/num/str/None)."""
    if v is None or isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return float(v)
    return str(v).strip()


def generate_all_answer_keys(rulebook: dict) -> dict:
    """
    Generate answer keys directly from the rulebook's seed data.

    The rulebook contains all data including computed field values.
    This function converts PascalCase field names to snake_case and
    exports to answer-keys/{entity}.json.

    Returns dict of entity_name -> list of records.

    NOTE: This function does NOT use a database. The rulebook is the
    single source of truth for answer keys. All execution substrates
    (Python, Go, Postgres, etc.) are tested equally against these keys.
    """
    print("Step 1: Generating answer keys from rulebook...", flush=True)

    # Clear and recreate answer-keys directory to remove stale files
    import shutil
    if os.path.exists(ANSWER_KEYS_DIR):
        shutil.rmtree(ANSWER_KEYS_DIR)
    os.makedirs(ANSWER_KEYS_DIR, exist_ok=True)

    entities = [e for e in discover_entities(rulebook) if e not in NON_GRADED_ENTITIES]
    skipped = sorted(NON_GRADED_ENTITIES)
    print(f"  Found {len(entities)} entities in rulebook"
          f" (excluding non-graded: {', '.join(skipped)})", flush=True)

    all_answer_keys = {}
    recompute_failures: list = []

    for entity_pascal in entities:
        entity_snake = to_snake_case(entity_pascal)
        pk = discover_primary_key(rulebook, entity_pascal)

        # Get seed data from rulebook
        raw_records = get_entity_data(rulebook, entity_pascal)

        if not raw_records:
            print(f"  -> {entity_snake}: No seed data (skipping)", flush=True)
            continue

        # Convert field names to snake_case
        records = [convert_record_to_snake_case(r) for r in raw_records]

        # Coerce values to match schema datatypes (e.g., "24" -> 24 for integers)
        schema = get_entity_schema(rulebook, entity_pascal)
        records = [coerce_record_types(r, schema) for r in records]

        # Recompute scalar `type=calculated` formulas from the current schema.
        # Cross-table fields (lookups / aggregations) are left to substrate-side
        # computation; only scalar calculated formulas are evaluated here.
        # Per-field eval errors accumulate in recompute_failures (see below).
        records = [
            _recompute_calculated_fields(r, schema, entity_pascal, recompute_failures)
            for r in records
        ]

        # Sort by primary key if available
        if pk and records and pk in records[0]:
            records = sorted(records, key=lambda r: r.get(pk, ''))

        # Save to file
        output_path = os.path.join(ANSWER_KEYS_DIR, f"{entity_snake}.json")
        with open(output_path, 'w') as f:
            json.dump(records, f, indent=2, default=str)

        all_answer_keys[entity_snake] = records
        print(f"  -> {entity_snake}: {len(records)} records", flush=True)

    if recompute_failures:
        # Persist the full list so generate-report.py can surface everything.
        failures_path = os.path.join(TESTING_DIR, "_recompute_failures.json")
        with open(failures_path, 'w') as f:
            json.dump(recompute_failures, f, indent=2, default=str)

        # Two severities, two policies:
        #   • missing-value — the Python engine couldn't evaluate AND there is no
        #     substrate-adopted value to fall back on. The answer key would be
        #     blank: that's a real hole, so RAISE (no silent blank — CLAUDE.md
        #     "Avoid Silent Fallbacks"). The fix is to run regenerate-answer-keys
        #     so Postgres computes the value, or to implement the function.
        #   • drift — the stored (substrate) value and the Python recompute
        #     disagree. The substrate is the oracle, so the key is still correct;
        #     this is a WARNING that the seed is stale vs the current formula.
        missing = [f for f in recompute_failures if f.get('severity') == 'missing-value']
        drift = [f for f in recompute_failures if f.get('severity') == 'drift']

        if drift:
            print(f"  ⚠ {len(drift)} calc field(s) DRIFT from current formulas "
                  f"(stored substrate value kept; run regenerate-answer-keys.sh "
                  f"to refresh):", flush=True)
            for f in drift:
                print(f"      - {f['entity']}.{f['field']} "
                      f"(id={f['record_id']!r}): {f['error']}", flush=True)

        if missing:
            summary_lines = [
                f"Answer-key generation has {len(missing)} field(s) with NO "
                f"computable value (Python engine cannot evaluate and no substrate "
                f"value is stored):"
            ]
            for fail in missing:
                summary_lines.append(
                    f"  - {fail['entity']}.{fail['field']} "
                    f"(record_id={fail['record_id']!r}): {fail['error']}"
                )
                summary_lines.append(f"      formula: {fail['formula']}")
            summary_lines.append(
                "Fix: run postgres-bootstrap/regenerate-answer-keys.sh so Postgres "
                "computes these values into the rulebook, or implement the missing "
                "function in orchestration/formula_parser.py."
            )
            summary_lines.append(f"Full details: {failures_path}")
            raise RuntimeError("\n".join(summary_lines))

    return all_answer_keys


# Legacy function - kept for postgres substrate's take-test.py
def generate_answer_keys_from_postgres(conn, rulebook: dict) -> dict:
    """
    Query all vw_* views and return as dict of entity_name -> records.

    NOTE: This function is used by the postgres substrate's take-test.py
    to query computed values from views. It is NOT used for generating
    the canonical answer keys (use generate_all_answer_keys() for that).
    """
    views = discover_views(conn)
    all_results = {}

    for view in views:
        entity = view_to_entity_name(view)
        pk = discover_primary_key(rulebook, entity)

        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Order by primary key if available
        if pk:
            cur.execute(f"SELECT * FROM {view} ORDER BY {pk}")
        else:
            cur.execute(f"SELECT * FROM {view}")

        rows = cur.fetchall()
        records = [dict(row) for row in rows]
        cur.close()

        all_results[entity] = records

    return all_results


# =============================================================================
# STEP 2: Generate Blank Tests (null computed columns)
# =============================================================================

def generate_all_blank_tests(all_answer_keys: dict, rulebook: dict) -> dict:
    """
    Create blank tests by nulling computed columns for each entity.
    Generates testing/blank-tests/{entity}.json files and testing/_metadata.json.
    """
    print("Step 2: Generating blank tests...", flush=True)

    # Clear and recreate blank-tests directory to remove stale files
    import shutil
    if os.path.exists(BLANK_TESTS_DIR):
        shutil.rmtree(BLANK_TESTS_DIR)
    os.makedirs(BLANK_TESTS_DIR, exist_ok=True)

    all_blank_tests = {}
    entity_metadata = {}

    for entity, records in all_answer_keys.items():
        computed_cols = discover_computed_columns(rulebook, entity)
        pk = discover_primary_key(rulebook, entity)

        if not computed_cols:
            print(f"  -> {entity}: No computed columns (skipping blank test)", flush=True)
            continue

        # Create blank test by nulling computed columns
        blank_records = []
        for record in records:
            blank_record = dict(record)
            for col in computed_cols:
                if col in blank_record:
                    blank_record[col] = None
            blank_records.append(blank_record)

        # Save to file
        output_path = os.path.join(BLANK_TESTS_DIR, f"{entity}.json")
        with open(output_path, 'w') as f:
            json.dump(blank_records, f, indent=2, default=str)

        all_blank_tests[entity] = blank_records

        # Track metadata for grading context
        entity_metadata[entity] = {
            "primary_key": pk,
            "computed_columns": computed_cols,
            "record_count": len(blank_records)
        }

        print(f"  -> {entity}: Nulled {len(computed_cols)} columns: {', '.join(computed_cols)}", flush=True)

    # Generate shared entity metadata at testing/_metadata.json (not in blank-tests/)
    metadata_path = os.path.join(TESTING_DIR, "_metadata.json")
    with open(metadata_path, 'w') as f:
        json.dump(entity_metadata, f, indent=2)
    print(f"  -> Entity metadata: {len(entity_metadata)} entities -> testing/_metadata.json", flush=True)

    return all_blank_tests


# =============================================================================
# STEP 3: Run Substrate Tests
# =============================================================================

def get_substrates() -> list:
    """Return the substrates this project actually exercises.

    Scoping rule (mirrors orchestrate.sh::get_valid_substrates and
    generate-report.py::get_substrates): intersect the substrates declared
    by the active project's effortless.json ProjectTranspilers with the
    substrate directories that exist on disk, then place them in
    SUBSTRATE_ORDER. Falls back to "everything on disk" only when the
    project has no effortless.json (legacy / un-initialized projects).
    """
    if not os.path.isdir(SUBSTRATES_DIR):
        return []

    discovered = {
        name for name in os.listdir(SUBSTRATES_DIR)
        if not name.startswith('.')
        and os.path.isdir(os.path.join(SUBSTRATES_DIR, name))
    }

    from shared import get_active_project_substrates
    domain = os.path.basename(DOMAIN_DIR.rstrip(os.sep))

    # If the active domain has no effortless.json, self-heal by running
    # `effortless -init` in the domain dir, then re-read. Reuse the already
    # correctly-resolved DOMAIN_DIR (rulebook-examples/ or toy-rulebooks/)
    # instead of recomputing it — see _find_domain_dir() above.
    domain_dir_abs = DOMAIN_DIR
    effortless_json = os.path.join(domain_dir_abs, "effortless.json")
    if not os.path.exists(effortless_json):
        print(
            f"effortless.json missing in {domain_dir_abs} — running "
            f"'effortless -init' to initialize the project...",
            flush=True,
        )
        subprocess.run(["effortless", "-init"], cwd=domain_dir_abs, check=True)

    declared = get_active_project_substrates(domain)  # raises if still missing
    if not declared:
        raise RuntimeError(
            f"effortless.json in {domain_dir_abs} has no enabled "
            "ProjectTranspilers. Add transpilers and rebuild."
        )

    allowed = [s for s in declared if s in discovered]
    allowed_set = set(allowed)
    ordered = [n for n in SUBSTRATE_ORDER if n in allowed_set]
    tail = sorted(allowed_set - set(SUBSTRATE_ORDER))
    return ordered + tail


def prepare_substrate_for_test(substrate_name: str) -> int:
    """
    Prepare a substrate for testing by clearing its test-answers/ directory.
    Substrates read blank-tests from the domain's testing/blank-tests/ location
    and write answers to the domain's testing/<substrate>/test-answers/.
    Returns number of test files that will be processed.
    """
    import shutil

    substrate_test_answers = get_substrate_test_answers_dir(substrate_name)

    # Clear and recreate test-answers directory
    if os.path.exists(substrate_test_answers):
        shutil.rmtree(substrate_test_answers)
    os.makedirs(substrate_test_answers, exist_ok=True)

    # Count blank test files (for reporting)
    count = 0
    for filename in os.listdir(BLANK_TESTS_DIR):
        if filename.endswith('.json') and not filename.startswith('_'):
            count += 1

    return count


def run_substrate_test(substrate_name: str) -> tuple:
    """Run a substrate's take-test.sh and return (success, error, elapsed, skipped)"""
    substrate_dir = os.path.join(SUBSTRATES_DIR, substrate_name)
    script_path = os.path.join(substrate_dir, "take-test.sh")

    if not os.path.exists(script_path):
        return False, "No take-test.sh found", 0.0, False

    start_time = time.time()
    try:
        result = subprocess.run(
            ["bash", script_path],
            cwd=substrate_dir,
            capture_output=True,
            text=True,
            timeout=600  # 10 minutes for LLM substrates like english
        )
        elapsed = time.time() - start_time

        # Check if substrate was skipped by user
        if "SUBSTRATE_SKIPPED" in result.stdout:
            return True, None, elapsed, True

        if result.returncode != 0:
            return False, f"Script failed: {result.stderr[:200]}", elapsed, False

        return True, None, elapsed, False

    except subprocess.TimeoutExpired:
        elapsed = time.time() - start_time
        return False, "Script timed out", elapsed, False
    except Exception as e:
        elapsed = time.time() - start_time
        return False, str(e), elapsed, False


def get_substrate_answers(substrate_name: str, rulebook: dict) -> dict:
    """
    Get all test-answers from a substrate's answer dir.
    Returns dict of entity_name -> list of records.

    If the substrate's test-answers/ dir does not exist, returns {} — that's a
    legitimate state (substrate was skipped, or crashed before writing). The
    grader records every expected (entity, field) pair as a miss in that case.

    If the dir exists but a JSON file inside is invalid, RAISE with the exact
    file path — corrupt output is a bug, not a "score it zero" situation.
    Substrates MUST write to the multi-entity test-answers/<entity>.json
    layout; there is no legacy single-file path.
    """
    answers_dir = get_substrate_test_answers_dir(substrate_name)

    if not os.path.isdir(answers_dir):
        return {}

    answers: dict = {}
    for file in sorted(glob.glob(os.path.join(answers_dir, "*.json"))):
        entity = os.path.basename(file).replace('.json', '')
        with open(file, 'r') as f:
            try:
                answers[entity] = json.load(f)
            except json.JSONDecodeError as e:
                raise json.JSONDecodeError(
                    f"Invalid JSON in {file}: {e.msg}",
                    e.doc, e.pos,
                ) from e
    return answers


# =============================================================================
# STEP 4: Grade Substrates (Generic Field-by-Field Comparison)
# =============================================================================

_NULL_STRINGS = {"", "none", "null", "n/a", "na"}
# Across substrates, the empty/absent/zero/false slot gets serialized many ways
# (Python None, SQL NULL, CSV empty string, "0", 0, False, "false", "no", etc).
# For grading we treat all of them as one equivalence class — substrate-level
# representational drift should never fail a conformance test on its own.
_FALSY_STRINGS = _NULL_STRINGS | {"0", "0.0", "false", "f", "no"}


def _is_nullish(val) -> bool:
    if val is None:
        return True
    if isinstance(val, bool):
        return val is False
    if isinstance(val, (int, float)):
        return val == 0
    if isinstance(val, str) and val.strip().lower() in _FALSY_STRINGS:
        return True
    return False


def _try_number(val):
    """Return val as float if it looks like a number, else None.

    Handles ints, floats, numeric strings with optional whitespace and commas
    (e.g. "1,000.0"). Returns None for booleans on purpose — bool comparison
    is handled separately so True/1 don't accidentally collide with 1.0.
    """
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip().replace(",", "")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _normalize_string(val) -> str:
    """Trim, collapse internal whitespace, and lowercase a string value."""
    if val is None:
        return ""
    s = str(val).strip()
    # Collapse any run of whitespace to a single space so " a  b " == "a b".
    s = " ".join(s.split())
    return s.lower()


def compare_values(expected, actual, datatype: str = None) -> bool:
    """Compare two values forgivingly across the type-representation drift that
    naturally occurs between substrates (Python None vs SQL NULL vs CSV empty,
    int vs float, "1,000" vs 1000, " Internal " vs "internal", etc.).

    Order of comparison:
      1. Both null-ish (None, "", "null", "none", "n/a") → equal.
      2. Datatype-aware compare when datatype is provided
         (boolean, integer/float/number/decimal, date/datetime).
      3. Numeric-aware compare when BOTH values look numeric — within a small
         epsilon, so 1 == 1.0 == "1" == "1.00".
      4. List/tuple compare element-wise (order-insensitive — lookup multi-values
         commonly come back in any order).
      5. Otherwise, normalized-string compare (trim, collapse whitespace,
         case-insensitive).

    This is the conservative variant — it's still deterministic. Fuzzy/semantic
    grading is a separate tool (llm-fuzzy-grader.py).
    """
    if _is_nullish(expected) and _is_nullish(actual):
        return True
    # One side null-ish, the other not, is a real disagreement.
    if _is_nullish(expected) != _is_nullish(actual):
        # Boolean is a special case: a null actual canonicalizes to False, which
        # may legitimately equal a False expected — fall through to the boolean
        # branch below for that one.
        if (datatype or "").lower() != "boolean":
            return False

    dt = (datatype or "").lower()

    # ---- Booleans -----------------------------------------------------------
    def normalize_bool(val):
        if val is None or (isinstance(val, str) and val.strip() == ""):
            return False
        if isinstance(val, bool):
            return val
        if isinstance(val, (int, float)):
            return bool(val)
        s = str(val).strip().lower()
        if s in ("true", "1", "yes", "y", "t"):
            return True
        if s in ("false", "0", "no", "n", "f", "none", "null"):
            return False
        return val  # unknown — let it fall through to equality

    if dt == "boolean":
        return normalize_bool(expected) == normalize_bool(actual)

    # ---- Numbers ------------------------------------------------------------
    NUMERIC_DTS = {"integer", "int", "float", "number", "numeric", "decimal", "currency"}
    if dt in NUMERIC_DTS:
        e_num, a_num = _try_number(expected), _try_number(actual)
        if e_num is not None and a_num is not None:
            return abs(e_num - a_num) < 1e-9
        # Datatype said numeric but one side wasn't parseable — fall through
        # to string compare so a clearly-wrong text answer still fails.

    # ---- Dates / datetimes --------------------------------------------------
    if dt in {"date", "datetime", "timestamp"}:
        e_norm = _normalize_string(expected).replace("t", " ").split(".")[0].rstrip("z")
        a_norm = _normalize_string(actual).replace("t", " ").split(".")[0].rstrip("z")
        # Trim trailing zero time so "2026-05-12" == "2026-05-12 00:00:00".
        for suffix in (" 00:00:00", " 00:00"):
            if e_norm.endswith(suffix):
                e_norm = e_norm[: -len(suffix)].rstrip()
            if a_norm.endswith(suffix):
                a_norm = a_norm[: -len(suffix)].rstrip()
        if e_norm == a_norm:
            return True
        # Fall through to give plain-string compare a chance too.

    # ---- Lists --------------------------------------------------------------
    if isinstance(expected, (list, tuple)) or isinstance(actual, (list, tuple)):
        e_list = list(expected) if isinstance(expected, (list, tuple)) else [expected]
        a_list = list(actual) if isinstance(actual, (list, tuple)) else [actual]
        if len(e_list) != len(a_list):
            return False
        # Order-insensitive: every expected element must have a match in actual.
        a_remaining = list(a_list)
        for e in e_list:
            for i, a in enumerate(a_remaining):
                if compare_values(e, a, None):
                    a_remaining.pop(i)
                    break
            else:
                return False
        return True

    # ---- Generic numeric-aware compare --------------------------------------
    # Even when datatype wasn't provided or wasn't numeric, if both sides
    # happen to look numeric, compare numerically. Catches "3" vs 3 and 1.0 vs 1.
    e_num, a_num = _try_number(expected), _try_number(actual)
    if e_num is not None and a_num is not None:
        return abs(e_num - a_num) < 1e-9

    # ---- String compare (trim, collapse whitespace, case-insensitive) -------
    return _normalize_string(expected) == _normalize_string(actual)


def grade_substrate(substrate_name: str, all_answer_keys: dict, rulebook: dict) -> dict:
    """
    Grade a substrate's answers against the answer keys.

    Hardened on 2026-05-12 per CONFORMANCE_CLEANUP_PLAN.md Step 4. Three
    behavioural rules now apply:

    1. The denominator is every (record x computed field) pair in the answer
       keys. We iterate over the answer keys, not over what the substrate
       happened to submit.
    2. Missing entity files, missing records, or null actuals where the
       answer key has a real value all count as FAILURES (not skips).
       Previously an entity where every actual was None was silently
       dropped from the totals — that is the cheat we are eliminating.
    3. The result breaks scores out by field class:
       scalar (calculated) vs lookup vs aggregation. Reports can present
       these independently so a substrate that does scalar math natively
       but cannot do cross-table joins reads as partial-honest, not 0/100.
    """
    results = {
        "substrate": substrate_name,
        "entities": {},
        "total_fields_tested": 0,
        "fields_passed": 0,
        "fields_failed": 0,
        # Per-field-class breakdown (Step 4 of cleanup plan).
        "by_class": {
            "calculated":  {"tested": 0, "passed": 0, "failed": 0},
            "lookup":      {"tested": 0, "passed": 0, "failed": 0},
            "aggregation": {"tested": 0, "passed": 0, "failed": 0},
        },
        "error": None,
        "elapsed_seconds": 0.0
    }

    substrate_answers = get_substrate_answers(substrate_name, rulebook)
    if substrate_answers is None:
        substrate_answers = {}

    # Iterate over the ANSWER KEYS, not the substrate's submissions. That way a
    # substrate that didn't write a test-answers/{entity}.json file gets scored
    # as zero on every (record x computed field) pair for that entity, instead
    # of silently disappearing from the denominator.
    for entity, answer_key in all_answer_keys.items():
        computed_cols = discover_computed_columns(rulebook, entity)
        pk = discover_primary_key(rulebook, entity)

        if not computed_cols:
            continue  # No computed fields to test on this entity.

        # Map column name -> (field_type, datatype) so we can both classify
        # (scalar / lookup / aggregation) and apply boolean-aware comparison.
        col_types = {}
        col_datatypes = {}
        for field in get_entity_schema(rulebook, entity):
            snake = to_snake_case(field.get('name', ''))
            col_types[snake] = field.get('type', 'calculated')
            col_datatypes[snake] = field.get('datatype', 'string')

        test_records = substrate_answers.get(entity, []) or []
        answers_by_pk = {}
        for record in test_records:
            pk_val = record.get(pk)
            if pk_val is not None:
                answers_by_pk[str(pk_val)] = record

        entity_result = {
            "total_records": len(answer_key),
            "computed_columns": computed_cols,
            "primary_key": pk,
            "fields_tested": 0,
            "fields_passed": 0,
            "fields_failed": 0,
            "by_class": {
                "calculated":  {"tested": 0, "passed": 0, "failed": 0},
                "lookup":      {"tested": 0, "passed": 0, "failed": 0},
                "aggregation": {"tested": 0, "passed": 0, "failed": 0},
            },
            "failures": [],
            "missing_file": entity not in substrate_answers,
        }

        for expected_record in answer_key:
            pk_val = str(expected_record.get(pk))
            actual_record = answers_by_pk.get(pk_val)  # may be None (missing record)
            record_present = actual_record is not None
            if actual_record is None:
                actual_record = {}

            for col in computed_cols:
                entity_result["fields_tested"] += 1
                results["total_fields_tested"] += 1
                field_class = col_types.get(col, 'calculated')
                if field_class not in entity_result["by_class"]:
                    field_class = 'calculated'

                entity_result["by_class"][field_class]["tested"] += 1
                results["by_class"][field_class]["tested"] += 1

                expected_val = expected_record.get(col)
                actual_val = actual_record.get(col)

                # If the expected value is null/empty, the substrate isn't on
                # the hook for it. Compare normally — compare_values will pass.
                passed = compare_values(expected_val, actual_val, col_datatypes.get(col))

                if passed:
                    entity_result["fields_passed"] += 1
                    entity_result["by_class"][field_class]["passed"] += 1
                    results["fields_passed"] += 1
                    results["by_class"][field_class]["passed"] += 1
                else:
                    entity_result["fields_failed"] += 1
                    entity_result["by_class"][field_class]["failed"] += 1
                    results["fields_failed"] += 1
                    results["by_class"][field_class]["failed"] += 1
                    entity_result["failures"].append({
                        "pk": pk_val,
                        "field": col,
                        "field_class": field_class,
                        "expected": expected_val,
                        "actual": actual_val,
                        "reason": (
                            "missing entity file" if entity_result["missing_file"]
                            else "missing record" if not record_present
                            else "wrong/null value"
                        ),
                    })

        results["entities"][entity] = entity_result

    if not substrate_answers and results["total_fields_tested"] == 0:
        # No answer keys had computed fields either; preserve old error signal.
        results["error"] = "No test-answers found"

    return results


def run_and_grade_all_substrates(all_answer_keys: dict, rulebook: dict) -> dict:
    """Run and grade each substrate"""
    print("Step 3: Running and grading tests for each substrate...", flush=True)
    print(flush=True)

    substrates = get_substrates()
    print(f"  Found {len(substrates)} substrates: {', '.join(substrates)}", flush=True)
    print(flush=True)

    all_grades = {}

    for i, substrate in enumerate(substrates, 1):
        print(f"  [{i}/{len(substrates)}] Testing {substrate}...", flush=True)

        # Prepare substrate for test (clear old test-answers)
        test_count = prepare_substrate_for_test(substrate)
        print(f"      Prepared {substrate} ({test_count} tests from shared testing/blank-tests/)", flush=True)

        # Run the test
        success, error, elapsed, skipped = run_substrate_test(substrate)

        # Grade the results
        grades = grade_substrate(substrate, all_answer_keys, rulebook)
        if error:
            grades["error"] = error

        # If substrate was skipped, use previous timing from metadata
        if skipped:
            print(f"      {substrate} was SKIPPED - using previous timing", flush=True)
            metadata = load_run_metadata(substrate)
            last_run = metadata.get("last_run", {})
            prev_elapsed = last_run.get("duration_seconds", 0.0)
            grades["elapsed_seconds"] = prev_elapsed
            grades["skipped"] = True
        else:
            grades["elapsed_seconds"] = elapsed
            grades["skipped"] = False

        all_grades[substrate] = grades

        # Generate report and print summary
        generate_substrate_report(substrate, grades, rulebook)
        print_substrate_test_summary(substrate, grades, rulebook)
        print("\n" * 5, flush=True)

    prune_stale_central_results(substrates)

    return all_grades


def prune_stale_central_results(active_substrates: list):
    """Drop _substrate_results.json entries for substrates no longer on disk.

    update_run_metadata()/save_run_metadata() only ever add or update one
    substrate at a time, so a removed substrate (e.g. a deleted
    execution-substrates/<name>/ dir) would otherwise leave its last score
    behind forever, silently misreporting it as still passing.
    """
    central = load_central_results()
    stale = [name for name in central if name not in active_substrates]
    if not stale:
        return
    for name in stale:
        del central[name]
    save_central_results(central)
    print(f"  Pruned stale substrate results: {', '.join(stale)}", flush=True)


# =============================================================================
# Reporting Functions
# =============================================================================

def format_duration(seconds: float) -> str:
    """Format duration in human-readable form.

    Uses whole seconds with '< 1s' for sub-second times to reduce git diff noise.
    """
    if seconds < 1:
        return "< 1s"
    elif seconds < 60:
        return f"{int(round(seconds))}s"
    else:
        mins = int(seconds // 60)
        secs = int(round(seconds % 60))
        return f"{mins}m {secs}s"


def generate_substrate_report(substrate_name: str, results: dict, rulebook: dict):
    """Generate test-results.md for a substrate (written into the domain testing folder)"""
    substrate_testing_dir = os.path.join(TESTING_DIR, substrate_name)
    os.makedirs(substrate_testing_dir, exist_ok=True)
    report_path = os.path.join(substrate_testing_dir, "test-results.md")

    total = results["total_fields_tested"]
    passed = results["fields_passed"]
    failed = results["fields_failed"]
    score = (passed / total * 100) if total > 0 else 0
    elapsed = results.get("elapsed_seconds", 0.0)

    # Check run metadata for error banner
    run_metadata = load_run_metadata(substrate_name)
    last_run = run_metadata.get("last_run", {})
    last_success = run_metadata.get("last_successful_run", {})

    lines = [
        f"# Test Results: {substrate_name}",
        "",
    ]

    # Add error banner if latest run failed but we have prior successful results
    if last_run.get("status") == "failure" and last_success:
        error_msg = last_run.get("error_message", "Unknown error")
        failure_time = last_run.get("timestamp", "Unknown")[:19].replace("T", " ")
        success_time = last_success.get("timestamp", "Unknown")[:19].replace("T", " ")
        success_results = last_success.get("test_results", {})

        lines.extend([
            "> **WARNING: Latest Run Failed**",
            "> ",
            f"> The test run at {failure_time} failed: `{error_msg}`",
            "> ",
            f"> Showing results from last successful run ({success_time}).",
            "",
        ])

        # Use the successful run's results for display
        if success_results:
            total = success_results.get("total_fields_tested", total)
            passed = success_results.get("fields_passed", passed)
            failed = success_results.get("fields_failed", failed)
            score = success_results.get("score", score)
            elapsed = last_success.get("duration_seconds", elapsed)

    lines.extend([
        "## Summary",
        "",
        "| Metric | Value |",
        "|--------|-------|",
        f"| Total Fields Tested | {total} |",
        f"| Passed | {passed} |",
        f"| Failed | {failed} |",
        f"| Score | {score:.1f}% |",
        f"| Duration | {format_duration(elapsed)} |",
        "",
    ])

    # By-field-class breakdown (Step 5 of CONFORMANCE_CLEANUP_PLAN.md).
    # Scalar (calculated) is what the substrate's native engine alone proves;
    # lookup and aggregation require cross-table semantics that several
    # substrates intentionally do not implement. Breaking these out lets the
    # scoreboard show partial-honest substrates as such rather than collapsing
    # everything into a single misleading percentage.
    by_class = results.get("by_class") or {}
    if any(c.get("tested", 0) for c in by_class.values()):
        lines.extend([
            "## Score by Field Class",
            "",
            "| Class | Passed | Tested | Score |",
            "|-------|--------|--------|-------|",
        ])
        for cls_name, label in (
            ("calculated",  "Scalar (calculated)"),
            ("lookup",      "Lookup (INDEX/MATCH)"),
            ("aggregation", "Aggregation (COUNTIFS/SUMIFS)"),
        ):
            c = by_class.get(cls_name, {})
            t = c.get("tested", 0)
            p = c.get("passed", 0)
            if t == 0:
                lines.append(f"| {label} | — | 0 | n/a |")
            else:
                lines.append(f"| {label} | {p} | {t} | {p / t * 100:.1f}% |")
        lines.append("")

    if results.get("error"):
        lines.extend([
            "## Error",
            "",
            "```",
            results["error"],
            "```",
            "",
        ])

    # Per-entity breakdown
    if results.get("entities"):
        lines.extend([
            "## Results by Entity",
            "",
        ])

        for entity, entity_result in results["entities"].items():
            e_total = entity_result["fields_tested"]
            e_passed = entity_result["fields_passed"]
            e_score = (e_passed / e_total * 100) if e_total > 0 else 0

            lines.extend([
                f"### {entity}",
                "",
                f"- Fields: {e_passed}/{e_total} ({e_score:.1f}%)",
                f"- Computed columns: {', '.join(entity_result['computed_columns'])}",
                "",
            ])

            if entity_result["failures"]:
                lines.extend([
                    "| PK | Field | Expected | Actual |",
                    "|-----|-------|----------|--------|",
                ])
                for failure in entity_result["failures"][:20]:
                    pk = failure["pk"]
                    field = failure["field"]
                    expected = str(failure["expected"])[:30]
                    actual = str(failure["actual"])[:30]
                    lines.append(f"| {pk} | {field} | {expected} | {actual} |")

                if len(entity_result["failures"]) > 20:
                    lines.append(f"| ... | ... | ({len(entity_result['failures']) - 20} more) | ... |")
                lines.append("")

    with open(report_path, 'w') as f:
        f.write('\n'.join(lines))


def print_substrate_test_summary(substrate_name: str, grades: dict, rulebook: dict):
    """Print a per-test breakdown for a substrate to console"""
    total = grades["total_fields_tested"]
    passed = grades["fields_passed"]
    failed = grades["fields_failed"]
    execution_failed = grades.get("error") and total == 0
    score = (passed / total * 100) if total > 0 else 0
    elapsed = grades.get("elapsed_seconds", 0.0)

    # Determine status
    if execution_failed:
        status_plain = "FAILED TO COMPUTE"
    elif grades.get("error"):
        status_plain = "ERROR"
    elif failed == 0:
        status_plain = "PASS"
    else:
        status_plain = "FAIL"

    score_color = get_score_color(score)

    box_width = 60
    if execution_failed:
        header_bg = RED_BG
        header_text = WHITE_TEXT
    else:
        header_bg = SKY_BLUE_BG
        header_text = DARK_TEXT

    print(f"  {header_bg}{header_text}┌{'─' * box_width}┐{RESET}", flush=True)
    print(f"  {header_bg}{header_text}│{BOLD} {substrate_name.upper():^{box_width - 2}} {RESET}{header_bg}{header_text}│{RESET}", flush=True)

    duration_str = format_duration(elapsed)
    if execution_failed:
        score_text = f"Score: --/-- (--%) - {status_plain}"
        print(f"  {header_bg}{header_text}│ {RED}{BOLD}{score_text:^{box_width - 2}}{RESET}{header_bg}{header_text} │{RESET}", flush=True)
    else:
        score_text = f"Score: {passed}/{total} ({score:.1f}%) - {status_plain}"
        print(f"  {header_bg}{header_text}│ {score_color}{BOLD}{score_text:^{box_width - 2}}{RESET}{header_bg}{header_text} │{RESET}", flush=True)

    duration_text = f"Duration: {duration_str}"
    print(f"  {header_bg}{header_text}│ {duration_text:^{box_width - 2}} │{RESET}", flush=True)
    print(f"  {header_bg}{header_text}├{'─' * box_width}┤{RESET}", flush=True)

    # Print per-entity results
    for entity, entity_result in grades.get("entities", {}).items():
        e_total = entity_result["fields_tested"]
        e_passed = entity_result["fields_passed"]
        e_failed = entity_result["fields_failed"]
        e_score = (e_passed / e_total * 100) if e_total > 0 else 0

        if e_failed == 0:
            row_bg = GREEN_BG
            icon = "✓"
            result_str = "PASS"
            text_color = GREEN
        else:
            row_bg = RED_BG
            icon = "✗"
            result_str = f"FAIL ({e_failed}/{e_total})"
            text_color = RED

        entity_display = entity[:35] if len(entity) > 35 else entity
        row_content = f"  {icon} {entity_display:<40} {result_str:>12} "
        print(f"  {row_bg}{WHITE_TEXT}│{row_content}│{RESET}", flush=True)

        # Show per-column breakdown
        failures_by_col = {}
        for f in entity_result.get("failures", []):
            col = f["field"]
            failures_by_col[col] = failures_by_col.get(col, 0) + 1

        for col in entity_result["computed_columns"]:
            col_failures = failures_by_col.get(col, 0)
            if col_failures == 0:
                col_status = f"{GREEN}✓{RESET}"
            else:
                col_status = f"{RED}{col_failures}{RESET}"
            col_display = col[:30] if len(col) > 30 else col
            print(f"  {header_bg}{header_text}│     {col_display:<40} {col_status:>18} │{RESET}", flush=True)

    if execution_failed:
        error_msg = grades.get("error", "Unknown error")[:50]
        print(f"  {RED_BG}{WHITE_TEXT}│ ERROR: {error_msg:<{box_width - 9}} │{RESET}", flush=True)

    print(f"  {header_bg}{header_text}└{'─' * box_width}┘{RESET}", flush=True)
    print(flush=True)


def generate_summary_report(all_grades: dict, rulebook: dict):
    """Generate all-tests-results.md with summary of all substrates"""
    print("Step 4: Generating summary report...", flush=True)

    # Collect all computed columns across all entities
    all_computed_cols = set()
    for substrate_grades in all_grades.values():
        for entity_result in substrate_grades.get("entities", {}).values():
            all_computed_cols.update(entity_result.get("computed_columns", []))

    lines = [
        "# Test Orchestrator Results",
        "",
        "## Configuration",
        "",
        f"- **Rulebook:** `{RULEBOOK_PATH}`",
        f"- **Substrates Tested:** {len(all_grades)}",
        f"- **Computed Columns Tested:** {len(all_computed_cols)}",
        "",
        "## Summary by Substrate",
        "",
        "| Substrate | Passed | Failed | Total | Score | Duration | Status |",
        "|-----------|--------|--------|-------|-------|----------|--------|",
    ]

    total_passed = 0
    total_failed = 0
    total_tests = 0
    total_time = 0.0

    # Sort by: 1) 100% substrates first (sorted by time), 2) <100% substrates at bottom (sorted by score desc)
    # Use substrate name as final tiebreaker for deterministic output
    # Round elapsed times to 0.5s buckets to prevent jitter from causing re-ordering
    def sort_key(name):
        g = all_grades[name]
        elapsed = g.get("elapsed_seconds", 0.0)
        # Round to 0.5s buckets for sorting stability
        elapsed_bucket = round(elapsed * 2) / 2
        p = g["fields_passed"]
        t = g["total_fields_tested"]
        score = (p / t * 100) if t > 0 else 0
        is_perfect = score >= 100.0
        # Primary: is_perfect (True=0, False=1 - so perfect scores come first)
        # Secondary: time bucket for perfect scores, -score for imperfect (highest score first among failures)
        # Tertiary: name for deterministic ordering when times are in same bucket
        return (0 if is_perfect else 1, elapsed_bucket if is_perfect else -score, name)

    for substrate_name in sorted(all_grades.keys(), key=sort_key):
        grades = all_grades[substrate_name]

        passed = grades["fields_passed"]
        failed = grades["fields_failed"]
        total = grades["total_fields_tested"]
        score = (passed / total * 100) if total > 0 else 0
        elapsed = grades.get("elapsed_seconds", 0.0)

        total_passed += passed
        total_failed += failed
        total_tests += total
        total_time += elapsed

        if grades.get("error"):
            status = f"ERROR: {grades['error'][:30]}"
        elif failed == 0:
            status = "PASS"
        else:
            status = "FAIL"

        lines.append(f"| {substrate_name} | {passed} | {failed} | {total} | {score:.1f}% | {format_duration(elapsed)} | {status} |")

    overall_score = (total_passed / total_tests * 100) if total_tests > 0 else 0

    lines.extend([
        "",
        "## Overall Statistics",
        "",
        "| Metric | Value |",
        "|--------|-------|",
        f"| Total Substrates | {len(all_grades)} |",
        f"| Total Fields Tested | {total_tests} |",
        f"| Total Passed | {total_passed} |",
        f"| Total Failed | {total_failed} |",
        f"| Overall Score | {overall_score:.1f}% |",
        f"| Total Duration | {format_duration(total_time)} |",
        "",
        "---",
        "",
        "*Generated by test-orchestrator.py (generic/domain-agnostic)*",
    ])

    with open(SUMMARY_PATH, 'w') as f:
        f.write('\n'.join(lines))

    print(f"  -> Summary written to {SUMMARY_PATH}", flush=True)
    print(f"  -> Overall: {total_passed}/{total_tests} ({overall_score:.1f}%)", flush=True)

    return total_passed, total_failed, total_tests, overall_score


def print_final_summary_table(all_grades: dict, rulebook: dict):
    """Print a final summary table to console"""
    print(flush=True)
    print("=" * 80, flush=True)
    print(f"{BOLD}FINAL RESULTS SUMMARY{RESET}", flush=True)
    print("=" * 80, flush=True)
    print(flush=True)

    # Header
    print(f"{'Substrate':<20} {'Passed':>8} {'Failed':>8} {'Total':>8} {'Score':>8} {'Duration':>10} {'Status':>12}", flush=True)
    print("-" * 80, flush=True)

    total_passed = 0
    total_failed = 0
    total_time = 0.0

    # Sort by: 1) 100% substrates first (sorted by time), 2) <100% substrates at bottom (sorted by score desc)
    # Use substrate name as final tiebreaker for deterministic output
    # Round elapsed times to 0.5s buckets to prevent jitter from causing re-ordering
    def sort_key(name):
        g = all_grades[name]
        elapsed = g.get("elapsed_seconds", 0.0)
        # Round to 0.5s buckets for sorting stability
        elapsed_bucket = round(elapsed * 2) / 2
        p = g["fields_passed"]
        t = g["total_fields_tested"]
        score = (p / t * 100) if t > 0 else 0
        is_perfect = score >= 100.0
        # Primary: is_perfect (True=0, False=1 - so perfect scores come first)
        # Secondary: time bucket for perfect scores, -score for imperfect (highest score first among failures)
        # Tertiary: name for deterministic ordering when times are in same bucket
        return (0 if is_perfect else 1, elapsed_bucket if is_perfect else -score, name)

    for substrate_name in sorted(all_grades.keys(), key=sort_key):
        grades = all_grades[substrate_name]
        passed = grades["fields_passed"]
        failed = grades["fields_failed"]
        total = grades["total_fields_tested"]
        score = (passed / total * 100) if total > 0 else 0
        elapsed = grades.get("elapsed_seconds", 0.0)

        total_passed += passed
        total_failed += failed
        total_time += elapsed

        score_color = get_score_color(score)

        if grades.get("error") and total == 0:
            status = f"{RED}FAILED{RESET}"
        elif failed == 0:
            status = f"{GREEN}PASS{RESET}"
        else:
            status = f"{YELLOW}PARTIAL{RESET}"

        print(f"{substrate_name:<20} {passed:>8} {failed:>8} {total:>8} {score_color}{score:>7.1f}%{RESET} {format_duration(elapsed):>10} {status:>12}", flush=True)

    print("-" * 80, flush=True)
    overall_total = total_passed + total_failed
    overall_score = (total_passed / overall_total * 100) if overall_total > 0 else 0
    print(f"{BOLD}{'OVERALL':<20}{RESET} {total_passed:>8} {total_failed:>8} {overall_total:>8} {BOLD}{overall_score:>7.1f}%{RESET} {format_duration(total_time):>10}", flush=True)
    print(flush=True)


# =============================================================================
# CLEANUP: Revert files with no REAL changes
# =============================================================================

def strip_timing_fields(obj, timing_fields):
    """Recursively remove timing-related fields from a JSON object."""
    if isinstance(obj, dict):
        return {k: strip_timing_fields(v, timing_fields) for k, v in obj.items()
                if k not in timing_fields}
    elif isinstance(obj, list):
        return [strip_timing_fields(item, timing_fields) for item in obj]
    return obj


def normalize_md_timing(content: str) -> str:
    """
    Normalize timing values and row order in markdown to enable comparison.

    Replaces duration values like "35s", "< 1s", "1m 23s" with a placeholder
    so that only-timing changes can be detected.

    Also sorts table rows by substrate name so that row order differences
    (caused by timing jitter affecting sort) don't register as real changes.
    """
    # Pattern for duration in table cells: | 35s | or | < 1s | or | 1m 23s |
    # Match patterns like: "35s", "< 1s", "1m 23s", "2m", etc.
    normalized = re.sub(r'\|\s*(<\s*)?\d+m?\s*\d*s?\s*\|', '| DURATION |', content)

    # Sort table rows alphabetically by first column (substrate name)
    # This ensures row order differences don't cause false positives
    lines = normalized.split('\n')
    result_lines = []
    table_rows = []
    in_table = False

    for line in lines:
        # Detect table data rows (start with | and have substrate names, not headers/separators)
        if line.startswith('|') and not line.startswith('|--') and not '| Substrate |' in line and not '| Metric |' in line:
            table_rows.append(line)
            in_table = True
        else:
            # If we were collecting table rows, sort and flush them
            if table_rows:
                table_rows.sort()
                result_lines.extend(table_rows)
                table_rows = []
            result_lines.append(line)
            in_table = False

    # Flush any remaining table rows
    if table_rows:
        table_rows.sort()
        result_lines.extend(table_rows)

    return '\n'.join(result_lines)


def revert_md_if_only_timing_changes(file_path: str) -> bool:
    """
    Check if the only changes in a markdown file are timing/duration values.
    If so, revert the file. Returns True if reverted, False otherwise.

    This is specifically for all-tests-results.md where Duration columns
    may change (e.g., "35s" -> "36s") without any real test result changes.
    """
    if not os.path.exists(file_path):
        return False

    try:
        rel_path = os.path.relpath(file_path, PROJECT_ROOT)

        # Check if file has uncommitted changes (staged or unstaged)
        result = subprocess.run(
            ['git', 'status', '--porcelain', '--', rel_path],
            capture_output=True, text=True, cwd=PROJECT_ROOT
        )
        if not result.stdout.strip():
            return False  # No changes

        # Load current file content
        with open(file_path, 'r') as f:
            current_content = f.read()

        # Get previous version from git
        result = subprocess.run(
            ['git', 'show', f'HEAD:{rel_path}'],
            capture_output=True, text=True, cwd=PROJECT_ROOT
        )
        if result.returncode != 0:
            return False  # File doesn't exist in previous commit

        previous_content = result.stdout

        # Normalize timing values in both versions
        current_normalized = normalize_md_timing(current_content)
        previous_normalized = normalize_md_timing(previous_content)

        # Compare
        if current_normalized != previous_normalized:
            # Real differences exist beyond timing values
            return False

        # All changes are timing-only - revert this file (both staged and unstaged)
        # First unstage if staged
        subprocess.run(
            ['git', 'reset', 'HEAD', '--', rel_path],
            capture_output=True, text=True, cwd=PROJECT_ROOT
        )
        # Then revert
        result = subprocess.run(
            ['git', 'checkout', '--', rel_path],
            capture_output=True, text=True, cwd=PROJECT_ROOT
        )

        if result.returncode == 0:
            print(f"  REVERTED (timing values only): {rel_path}", flush=True)
            return True
        return False

    except Exception as e:
        print(f"  WARNING: Error checking {file_path}: {e}", flush=True)
        return False


def revert_if_only_duration_changes(file_path: str) -> bool:
    """
    Check if the only changes in a JSON file are to timing-related fields.
    If so, revert the file. Returns True if reverted, False otherwise.

    Timing-related fields that are considered "noise" (not real changes):
    - duration_seconds: varies per run

    Algorithm:
    - Load current JSON and strip timing fields
    - Load previous commit JSON and strip timing fields
    - If they're identical after stripping, revert
    """
    # Fields that are timing-related metadata, not actual test results
    TIMING_FIELDS = {'duration_seconds'}

    if not os.path.exists(file_path):
        return False

    try:
        rel_path = os.path.relpath(file_path, PROJECT_ROOT)

        # Check if file has uncommitted changes (staged or unstaged)
        result = subprocess.run(
            ['git', 'status', '--porcelain', '--', rel_path],
            capture_output=True, text=True, cwd=PROJECT_ROOT
        )
        if not result.stdout.strip():
            return False  # No changes

        # Load current JSON
        with open(file_path, 'r') as f:
            current_data = json.load(f)

        # Get previous version from git
        result = subprocess.run(
            ['git', 'show', f'HEAD:{rel_path}'],
            capture_output=True, text=True, cwd=PROJECT_ROOT
        )
        if result.returncode != 0:
            return False  # File doesn't exist in previous commit

        previous_data = json.loads(result.stdout)

        # Strip timing fields from both
        current_stripped = strip_timing_fields(current_data, TIMING_FIELDS)
        previous_stripped = strip_timing_fields(previous_data, TIMING_FIELDS)

        # Compare
        if current_stripped != previous_stripped:
            # Real differences exist beyond timing fields
            return False

        # All changes are timing-only - revert this file (both staged and unstaged)
        # First unstage if staged
        subprocess.run(
            ['git', 'reset', 'HEAD', '--', rel_path],
            capture_output=True, text=True, cwd=PROJECT_ROOT
        )
        # Then revert
        result = subprocess.run(
            ['git', 'checkout', '--', rel_path],
            capture_output=True, text=True, cwd=PROJECT_ROOT
        )

        if result.returncode == 0:
            print(f"  REVERTED (timing fields only): {rel_path}", flush=True)
            return True
        return False

    except Exception as e:
        print(f"  WARNING: Error checking {file_path}: {e}", flush=True)
        return False


def xlsx_to_json(workbook) -> dict:
    """
    Export an xlsx workbook to a JSON structure containing all tabs,
    cells, values, and formulas.

    Returns a dict like:
    {
        "sheets": {
            "SheetName": {
                "dimensions": {"rows": 10, "cols": 5},
                "cells": [
                    {"row": 1, "col": 1, "value": "Header", "formula": null},
                    {"row": 2, "col": 1, "value": "=A1+B1", "formula": "=A1+B1"},
                    ...
                ]
            }
        }
    }
    """
    result = {"sheets": {}}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_data = {
            "dimensions": {
                "rows": sheet.max_row or 0,
                "cols": sheet.max_column or 0
            },
            "cells": []
        }

        # Export all cells with their values/formulas
        for row in range(1, (sheet.max_row or 0) + 1):
            for col in range(1, (sheet.max_column or 0) + 1):
                cell = sheet.cell(row=row, column=col)
                cell_value = cell.value

                # Determine if it's a formula
                is_formula = isinstance(cell_value, str) and cell_value.startswith('=')

                sheet_data["cells"].append({
                    "row": row,
                    "col": col,
                    "value": cell_value,
                    "formula": cell_value if is_formula else None
                })

        result["sheets"][sheet_name] = sheet_data

    return result


def revert_xlsx_if_unchanged(file_path: str) -> bool:
    """
    Check if an xlsx file has no real content changes.
    Exports both current and previous xlsx to JSON (all tabs + formulas)
    in a temporary folder and compares them.

    If content and formulas are identical, revert the file.
    Returns True if reverted, False otherwise.

    Algorithm:
    1. Export current xlsx to JSON (all tabs + formulas)
    2. Export previous commit's xlsx to JSON (all tabs + formulas)
    3. Compare the JSON structures
    4. If identical, revert the file
    5. Clean up temp files immediately
    """
    if not os.path.exists(file_path):
        return False

    try:
        import openpyxl
        import tempfile

        # Get relative path for git commands
        rel_path = os.path.relpath(file_path, PROJECT_ROOT)

        # Check if file has uncommitted changes (staged or unstaged)
        result = subprocess.run(
            ['git', 'status', '--porcelain', '--', rel_path],
            capture_output=True, text=True, cwd=PROJECT_ROOT
        )
        if not result.stdout.strip():
            return False  # No changes

        # Create temp directory for our work
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_xlsx_path = os.path.join(tmp_dir, 'previous.xlsx')
            current_json_path = os.path.join(tmp_dir, 'current.json')
            previous_json_path = os.path.join(tmp_dir, 'previous.json')

            # Load current workbook and export to JSON
            try:
                current_wb = openpyxl.load_workbook(file_path, data_only=False)
            except Exception:
                return False

            current_json = xlsx_to_json(current_wb)
            with open(current_json_path, 'w') as f:
                json.dump(current_json, f, sort_keys=True, default=str)

            # Get previous version from git
            result = subprocess.run(
                ['git', 'show', f'HEAD:{rel_path}'],
                capture_output=True, cwd=PROJECT_ROOT
            )

            if result.returncode != 0:
                return False  # File doesn't exist in previous commit

            with open(tmp_xlsx_path, 'wb') as f:
                f.write(result.stdout)

            # Load previous workbook and export to JSON
            previous_wb = openpyxl.load_workbook(tmp_xlsx_path, data_only=False)
            previous_json = xlsx_to_json(previous_wb)
            with open(previous_json_path, 'w') as f:
                json.dump(previous_json, f, sort_keys=True, default=str)

            # Compare JSON structures
            if current_json != previous_json:
                return False  # Real differences exist

            # All content and formulas are identical - revert
            # First unstage if staged
            subprocess.run(
                ['git', 'reset', 'HEAD', '--', rel_path],
                capture_output=True, text=True, cwd=PROJECT_ROOT
            )
            # Then revert
            result = subprocess.run(
                ['git', 'checkout', '--', rel_path],
                capture_output=True, text=True, cwd=PROJECT_ROOT
            )

            if result.returncode == 0:
                print(f"  REVERTED (no content/formula changes): {rel_path}", flush=True)
                return True
            return False

        # TemporaryDirectory is automatically cleaned up here

    except ImportError:
        print("  WARNING: openpyxl not available for xlsx comparison", flush=True)
        return False
    except Exception as e:
        print(f"  WARNING: Error comparing {file_path}: {e}", flush=True)
        return False


def cleanup_unchanged_files():
    """
    Revert files that have no REAL changes to reduce noise in commits.

    1. _substrate_results.json - revert if only duration_seconds changed
    2. all-tests-results.md - revert if only duration values changed (e.g., 35s -> 36s)
    3. Per-substrate test-results.md - revert if only duration changed
    4. *.xlsx files - revert if content and formulas are identical to previous commit
    """
    print("=" * 60, flush=True)
    print("CLEANUP: Reverting files with no real changes", flush=True)
    print("=" * 60, flush=True)

    reverted_count = 0

    # 1. Check _substrate_results.json
    if revert_if_only_duration_changes(CENTRAL_RESULTS_PATH):
        reverted_count += 1

    # 2. Check all-tests-results.md
    if revert_md_if_only_timing_changes(SUMMARY_PATH):
        reverted_count += 1

    # 3. Find and check all changed files (staged or unstaged)
    result = subprocess.run(
        ['git', 'status', '--porcelain'],
        capture_output=True, text=True, cwd=PROJECT_ROOT
    )

    if result.returncode == 0:
        for line in result.stdout.strip().split('\n'):
            if not line:
                continue
            # Format is: XY filename (where XY are status codes)
            rel_path = line[3:].strip()  # Skip status codes and space
            full_path = os.path.join(PROJECT_ROOT, rel_path)

            # Check xlsx files
            if rel_path.endswith('.xlsx'):
                if revert_xlsx_if_unchanged(full_path):
                    reverted_count += 1

            # Check per-substrate test-results.md files
            elif rel_path.endswith('test-results.md') and ('execution-substrates/' in rel_path or 'testing/' in rel_path):
                if revert_md_if_only_timing_changes(full_path):
                    reverted_count += 1

    if reverted_count == 0:
        print("  No files reverted (all changes are real)", flush=True)
    else:
        print(f"  Total files reverted: {reverted_count}", flush=True)

    print(flush=True)


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 60, flush=True)
    print("TEST ORCHESTRATOR (Generic / Domain-Agnostic)", flush=True)
    print("=" * 60, flush=True)
    print(flush=True)

    # Load rulebook - the single source of truth
    rulebook = load_rulebook()
    entities = discover_entities(rulebook)
    print(f"Discovered {len(entities)} entities in rulebook: {', '.join(entities)}", flush=True)
    print(flush=True)

    # Step 1: Generate answer keys directly from rulebook (no database needed)
    # The rulebook contains all data including expected computed field values.
    # All substrates (Python, Go, Postgres, etc.) are tested equally against these.
    all_answer_keys = generate_all_answer_keys(rulebook)
    print(flush=True)

    # Step 2: Generate blank tests
    generate_all_blank_tests(all_answer_keys, rulebook)
    print(flush=True)

    # Step 3: Run and grade each substrate
    all_grades = run_and_grade_all_substrates(all_answer_keys, rulebook)
    print(flush=True)

    # Step 4: Generate summary report
    print("\n" * 3, flush=True)
    generate_summary_report(all_grades, rulebook)
    print(flush=True)

    # Step 5: Print final summary table
    print_final_summary_table(all_grades, rulebook)

    # Step 6: Clean up files with no real changes
    cleanup_unchanged_files()

    print("=" * 60, flush=True)
    print("DONE", flush=True)
    print("=" * 60, flush=True)


if __name__ == "__main__":
    main()
